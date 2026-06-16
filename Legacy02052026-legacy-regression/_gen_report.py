import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import Counter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Not Required Files"

header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor="1F4E79")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin = Side(style="thin", color="AAAAAA")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

cat_fills = {
    "Scratch / Temp":                      PatternFill("solid", fgColor="FFD7D7"),
    "Draft / Experimental":                PatternFill("solid", fgColor="FFE8C0"),
    "Standalone Script (not in tests/)":   PatternFill("solid", fgColor="FFF2CC"),
    "Redundant / Superseded":              PatternFill("solid", fgColor="E2EFDA"),
    "Unused Helper Module":                PatternFill("solid", fgColor="DAE8FC"),
    "Duplicate in src/":                   PatternFill("solid", fgColor="E1D5E7"),
    "Broken / Empty Placeholder":          PatternFill("solid", fgColor="F8CECC"),
}

headers = ["#", "File Name", "Category", "Reason Not Required"]
ws.append(headers)
for col in range(1, 5):
    cell = ws.cell(1, col)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = border
ws.row_dimensions[1].height = 30

data = [
    (1,  "09082025.py",                                          "Scratch / Temp",                     "Date-named temp file. Contains comment 'no longer needed'. Duplicate login/cookie logic already in TestLoginLegacy.py and test_handleCookie.py."),
    (2,  "1.py",                                                 "Scratch / Temp",                     "Generic single-digit name. Duplicate zip binder logic superseded by binder_generate_as_zip.py, which is what tests actually import."),
    (3,  "123.py",                                               "Scratch / Temp",                     "Generic 3-digit name. Duplicate excel-from-options logic superseded by binder_generate_excel_from_options.py."),
    (4,  "2205.py",                                              "Scratch / Temp",                     "Date-named file. All functional code is inside a triple-quoted string (effectively commented out). Dead code."),
    (5,  "last_13082025.py",                                     "Scratch / Temp",                     "Date-named file. 10-K tearsheet generation logic identical to delete3.py and tenkmaybe_itsFinal.py. All three are duplicates."),
    (6,  "sdflsdafjalksfjalfjkldfsdflasdlkfaslfkaklfdmskmadsadsd.py", "Scratch / Temp",               "Random-character filename. Contains early draft of manage_download() fully superseded by binder_common.py."),
    (7,  "haunuman.py",                                          "Draft / Experimental",               "Non-descriptive name. Early draft of manage_download() utility. Functionality fully replaced by binder_common.py."),
    (8,  "kalki.py",                                             "Draft / Experimental",               "Non-descriptive name. Alternate download manager using xlwings. Superseded by binder_common.py."),
    (9,  "kalki2_option.py",                                     "Draft / Experimental",               "Non-descriptive name. Another download manager variant (option 2). Superseded by binder_common.py."),
    (10, "tenkmaybe_itsFinal.py",                                "Draft / Experimental",               "'maybe' in filename signals uncertainty. Duplicate 10-K script identical to delete3.py and last_13082025.py."),
    (11, "delete3.py",                                           "Draft / Experimental",               "Name implies a scratch/deletion script. Duplicate 10-K generation logic identical to tenkmaybe_itsFinal.py."),
    (12, "playwright_boilerplate.py",                            "Draft / Experimental",               "Contains comment 'will think about this import later'. Unfinished boilerplate not imported anywhere in the active test suite."),
    (13, "test.py",                                              "Standalone Script (not in tests/)",  "Generic name. Redefines handle_cookie_popup already in test_handleCookie.py. Not in tests/ folder so not discovered by pytest."),
    (14, "test_main.py",                                         "Standalone Script (not in tests/)",  "Only contains asyncio.run(login()). A one-liner scratch file, not a proper test."),
    (15, "test_ten_kreportModule.py",                            "Standalone Script (not in tests/)",  "Not in tests/ folder so not run by pytest. Duplicate 10-K script."),
    (16, "QuickReport.py",                                       "Standalone Script (not in tests/)",  "Redefines handle_cookie_popup locally instead of importing test_handleCookie.py. Not used by any test."),
    (17, "TearsheetReport.py",                                   "Standalone Script (not in tests/)",  "Redefines handle_cookie_popup locally. Standalone tearsheet script not imported by any test."),
    (18, "Login.py",                                             "Standalone Script (not in tests/)",  "Uses sync_playwright + pyautogui. Project uses async via TestLoginLegacy.py. Never imported by any test."),
    (19, "saveTemplateLegacy.py",                                "Standalone Script (not in tests/)",  "Standalone script with hardcoded credentials and its own browser launch. Not imported by any test."),
    (20, "run_all_scripts1.py",                                  "Standalone Script (not in tests/)",  "Hardcoded path pointing to a specific developer's machine (C:\\Users\\renny_kurian\\PycharmProjects\\...). Broken on any other machine."),
    (21, "navigate_binder_byurl.py",                             "Standalone Script (not in tests/)",  "One-off navigation helper with a bug (page context exits before use). Not imported by any test."),
    (22, "TestCreateBinder.py",                                  "Standalone Script (not in tests/)",  "Standalone binder creation script. Not imported by any test in tests/."),
    (23, "TestAddtoBinderFromTearsheet.py",                      "Standalone Script (not in tests/)",  "Standalone add-to-binder script. Not imported by any test in tests/."),
    (24, "renamebinder.py",                                      "Standalone Script (not in tests/)",  "Standalone binder rename script. Not used by any test."),
    (25, "deletebinder.py",                                      "Standalone Script (not in tests/)",  "Standalone binder delete script. Not used by any test."),
    (26, "UploadDocument_binder.py",                             "Standalone Script (not in tests/)",  "Standalone document upload script. Not imported by any test."),
    (27, "CIQ Report (Landscape).py",                            "Standalone Script (not in tests/)",  "Standalone CIQ landscape report script. Not part of the test suite."),
    (28, "Detailed Report (Excel).py",                           "Standalone Script (not in tests/)",  "Standalone detailed Excel report script. Not part of the test suite."),
    (29, "Detailed Report (PDF).py",                             "Standalone Script (not in tests/)",  "Standalone detailed PDF report script. Not part of the test suite."),
    (30, "Detailed Report (Word).py",                            "Standalone Script (not in tests/)",  "Standalone detailed Word report script. Not part of the test suite."),
    (31, "Summary Report (Excel).py",                            "Standalone Script (not in tests/)",  "Standalone summary Excel report script. Not part of the test suite."),
    (32, "Summary Report (PDF).py",                              "Standalone Script (not in tests/)",  "Standalone summary PDF report script. Not part of the test suite."),
    (33, "Summary Report (Word).py",                             "Standalone Script (not in tests/)",  "Standalone summary Word report script. Not part of the test suite."),
    (34, "Tearsheet Report (Word).py",                           "Standalone Script (not in tests/)",  "Standalone tearsheet Word report. Not part of the test suite."),
    (35, "TestCreateBinderoriginal.py",                          "Redundant / Superseded",             "'original' in filename signals it was replaced. Superseded by TestCreateBinder.py. Contains hardcoded credentials."),
    (36, "uploadDocument_binder1.py",                            "Redundant / Superseded",             "Numbered variant ('1') superseded by UploadDocument_binder.py. Uses its own browser launch without shared login."),
    (37, "binder_generate_zip_all_document.py",                  "Redundant / Superseded",             "NOT imported by tests (tests use binder_generate_as_zip.py). Uses old download_manager.py instead of binder_common.py."),
    (38, "download_manager.py",                                  "Unused Helper Module",               "Only imported by binder_generate_zip_all_document.py, which itself is not used by any test. Covered by binder_common.py."),
    (39, "download_mangerforBinder.py",                          "Unused Helper Module",               "Typo in name ('manger' instead of 'manager'). Duplicate download manager logic not imported by any test or active module."),
    (40, "downlodciqreport.py",                                  "Unused Helper Module",               "Typo in name ('downlod' instead of 'download'). Simple CIQ download helper not imported by any test."),
    (41, "ciq_download_manager.py",                              "Unused Helper Module",               "CIQ-specific download manager. Not imported by any active test file."),
    (42, "tenk_download_manager.py",                             "Unused Helper Module",               "Only imported by tenk_1808.py, which itself is not part of the test suite."),
    (43, "tenk_1808.py",                                         "Unused Helper Module",               "Date-named 10-K script not in tests/ and not imported by any test. References tenk_download_manager.py (also not needed)."),
    (44, "moduleCreateBinder.py",                                "Unused Helper Module",               "Contains comment 'use later when you want to import this module'. Not imported by any test."),
    (45, "src/binder_generate_pdf_all_document.py",              "Duplicate in src/",                  "Exact duplicate of the root-level file. Tests import from root; this src/ version is never referenced."),
    (46, "tests/test_pure_functions.py",                         "Broken / Empty Placeholder",         "References undefined function some_pure_function(). Will always fail with NameError. No real test logic."),
    (47, "tests/test_binder_integration.py",                     "Broken / Empty Placeholder",         "All 5 tests only contain 'assert True'. Pure placeholder with no actual test logic."),
]

for row in data:
    ws.append(row)
    r = ws.max_row
    fill = cat_fills.get(row[2], PatternFill("solid", fgColor="FFFFFF"))
    for col in range(1, 5):
        cell = ws.cell(r, col)
        cell.fill = fill
        cell.alignment = left if col in (2, 4) else center
        cell.border = border

ws.column_dimensions["A"].width = 5
ws.column_dimensions["B"].width = 55
ws.column_dimensions["C"].width = 32
ws.column_dimensions["D"].width = 72
ws.freeze_panes = "A2"

# Summary sheet
ws2 = wb.create_sheet("Summary")
s_headers = ["Category", "Count", "Color Key"]
ws2.append(s_headers)
for col in range(1, 4):
    cell = ws2.cell(1, col)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = border
ws2.row_dimensions[1].height = 25

counts = Counter(row[2] for row in data)
for cat, cnt in sorted(counts.items()):
    ws2.append([cat, cnt, ""])
    r = ws2.max_row
    fill = cat_fills.get(cat, PatternFill("solid", fgColor="FFFFFF"))
    for col in range(1, 4):
        cell = ws2.cell(r, col)
        cell.fill = fill
        cell.alignment = left if col == 1 else center
        cell.border = border

ws2.append(["TOTAL", sum(counts.values()), ""])
r = ws2.max_row
total_style = Font(bold=True, color="FFFFFF", size=11)
for col in range(1, 4):
    cell = ws2.cell(r, col)
    cell.font = total_style
    cell.fill = header_fill
    cell.alignment = center
    cell.border = border

ws2.column_dimensions["A"].width = 38
ws2.column_dimensions["B"].width = 10
ws2.column_dimensions["C"].width = 18
ws2.freeze_panes = "A2"

output = "not_required_files.xlsx"
wb.save(output)
print(f"Excel saved: {output}")
print(f"Total not-required files: {len(data)}")
